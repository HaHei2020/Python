#!/usr/bin/env python
# coding: UTF-8
# @Time    : 2017/11/17 11:25
# @Author  : HaHei
# 玩家留存人数、留存率， UID（账号）区分

import pymysql
from views import get_query_config
from models import get_db_config
import datetime

''' 选择 划分依据 '''
def get_retention():
    query_basis = get_query_config.getConfig("retenion", "query_basis")
    str_retenion_day = get_query_config.getConfig("retenion", "retenion_day")
    str_retenion_country = get_query_config.getConfig("retenion", "retenion_country")
    register_date_start = datetime.datetime.strptime(get_query_config.getConfig("retenion", "register_date_start"), '%Y-%m-%d')
    register_date_end = datetime.datetime.strptime(get_query_config.getConfig("retenion", "register_date_end"), '%Y-%m-%d')
    query_end_date = ''  # 暂废！
    #query_end_date = datetime.datetime.strptime(get_query_config.getConfig("retenion", "query_end_date"), '%Y-%m-%d')

    # if query_end_date < register_date_start or query_end_date < register_date_end:
    #     print("Error: 查询截止日期 不能小于 注册日期！")
    #     print("请重新配置！")
    #     return

    if register_date_start > register_date_end:
        print("Error: 注册查询开始日期 不能大于 注册查询结束日期！")
        print("请重新配置！")
        return

    retenion_day = []
    for d in str_retenion_day.split(', '):
        retenion_day.append(int(d))

    retenion_country = []
    if str_retenion_country != 'None':   # 区分 国家
        for c in str_retenion_country.split(', '):
            retenion_country.append(c)
        get_country_retenion(register_date_start, register_date_end, query_end_date, query_basis, retenion_day, retenion_country)

    else:
        get_retenion(register_date_start, register_date_end, query_end_date, query_basis, retenion_day)

''' 计算 整体 留存 '''
def get_retenion(register_date_start, register_date_end, query_end_date, query_basis, retenion_day):
    print("计算留存数据...")
    retenion_results = []
    register_date_days = str(register_date_end - register_date_start).split(' ')[0]  # 计算 注册天数 差值  如： 5 days
    for day in range(int(register_date_days)+1):
        register_date = register_date_start + datetime.timedelta(day)
        str_register_date = str(register_date).split(' ')[0] + '%'   # day 从 0 开始
        #print(register_date)     # 输出格式： '2017-07-07%'
        #print("注册日期：%s，留存数据如下：" %(str(register_date).split(' ')[0]))
        retenion_results.append({"register_date": str(register_date).split(' ')[0], "retenion_day": ""})

        retenion_info = []
        for d in retenion_day:  # 0：注册日期  算 注册人数 数组[]是：计算留存的天数，如：2留，3留等
            query_login_date = register_date + datetime.timedelta(d)
            #if query_login_date < query_end_date:  # 如果 小于 查询截止日期，可以继续查询
            str_query_login_date = str(query_login_date).split(' ')[0] + '%'

            # 留存 sql
            if query_basis == 'UID':
                query_sql = (
                    "SELECT PlayerRegister.UID AS 'UID', PlayerRegister.dtEventTime AS '注册时间', PlayerLogin.dtEventTime AS '登录时间' " \
                    "FROM PlayerLogin LEFT JOIN PlayerRegister " \
                    "ON PlayerRegister.UID = PlayerLogin.UID " \
                    "WHERE PlayerRegister.dtEventTime LIKE '%s' AND PlayerLogin.dtEventTime LIKE '%s' " \
                    "GROUP BY PlayerLogin.UID" %(str_register_date, str_query_login_date)
                )
            if query_basis == 'DVID':
                query_sql = (
                    "SELECT PlayerRegister.DVID AS 'DVID', PlayerRegister.dtEventTime AS '注册时间', PlayerLogin.dtEventTime AS '登录时间' " \
                    "FROM PlayerLogin LEFT JOIN PlayerRegister " \
                    "ON PlayerRegister.DVID = PlayerLogin.DVID " \
                    "WHERE PlayerRegister.dtEventTime LIKE '%s' AND PlayerLogin.dtEventTime LIKE '%s' " \
                    "GROUP BY PlayerLogin.DVID" %(str_register_date, str_query_login_date)
                )

            retenion_counts = get_mysql(query_sql)
            #print("%s日留存，人数为：%s" %(d+1, retenion_counts))
            retenion_info.append({"day": d+1, "count": retenion_counts})
        retenion_results[-1]['retenion_day'] = retenion_info
    print(retenion_results)

    write_excel(retenion_results, retenion_day)
    return



''' 分 国家 计算留存 '''
def get_country_retenion(register_date_start, register_date_end, query_end_date, query_basis, retenion_day, retenion_country):
    print("计算留存数据...")
    retenion_results = []
    register_date_days = str(register_date_end - register_date_start).split(' ')[0]  # 计算 注册天数 差值  如： 5 days
    for query_country in retenion_country:
        str_query_country = '%' + query_country + '%'
        #print("查询国家：", str_query_country)

        for day in range(int(register_date_days) + 1):
            register_date = register_date_start + datetime.timedelta(day)
            str_register_date = str(register_date).split(' ')[0] + '%'  # day 从 0 开始
            # print(register_date)     # 输出格式： '2017-07-07%'
            #print("注册日期：%s，留存数据如下：" % (str(register_date).split(' ')[0]))

            retenion_info = []
            for d in retenion_day:  # 0：注册日期  算 注册人数 数组[]是：计算留存的天数，如：2留，3留等
                query_login_date = register_date + datetime.timedelta(d)
                # if query_login_date < query_end_date:  # 如果 小于 查询截止日期，可以继续查询
                str_query_login_date = str(query_login_date).split(' ')[0] + '%'

                # 留存 sql
                if query_basis == 'UID':
                    query_sql = (
                        "SELECT PlayerRegister.UID AS 'UID', PlayerRegister.dtEventTime AS '注册时间', PlayerLogin.dtEventTime AS '登录时间', PlayerRegister.Country AS '国家' " \
                        "FROM PlayerLogin LEFT JOIN PlayerRegister " \
                        "ON PlayerRegister.UID = PlayerLogin.UID " \
                        "WHERE PlayerRegister.dtEventTime LIKE '%s' AND PlayerLogin.dtEventTime LIKE '%s' AND Country LIKE '%s' " \
                        "GROUP BY PlayerLogin.UID" % (str_register_date, str_query_login_date, str_query_country)
                    )
                if query_basis == 'DVID':
                    query_sql = (
                        "SELECT PlayerRegister.DVID AS 'DVID', PlayerRegister.dtEventTime AS '注册时间', PlayerLogin.dtEventTime AS '登录时间', PlayerRegister.Country AS '国家' " \
                        "FROM PlayerLogin LEFT JOIN PlayerRegister " \
                        "ON PlayerRegister.DVID = PlayerLogin.DVID " \
                        "WHERE PlayerRegister.dtEventTime LIKE '%s' AND PlayerLogin.dtEventTime LIKE '%s' AND Country LIKE '%s' " \
                        "GROUP BY PlayerLogin.DVID" % (str_register_date, str_query_login_date, str_query_country)
                    )

                retenion_counts = get_mysql(query_sql)
                #print("%s日留存，人数为：%s" % (d + 1, retenion_counts))
                retenion_info.append({"day": d + 1, "count": retenion_counts})
            retenion_results.append({"country": query_country, "register_date": str(register_date).split(' ')[0], "retenion_day": retenion_info})
        print(retenion_results)

    write_country_excel(retenion_results, retenion_day, retenion_country)
    return


''' 启动 mysql 连接 '''
def get_mysql(query_sql):
    # 获取 数据库 配置
    db_host = get_db_config.getConfig("database", "db_host")
    db_port = get_db_config.getConfig("database", "db_port")
    db_user = get_db_config.getConfig("database", "db_user")
    db_password = get_db_config.getConfig("database", "db_password")
    db_name = get_db_config.getConfig("database", "db_name")
    db_charset = get_db_config.getConfig("database", "db_charset")

    # 连接 数据库
    conn = pymysql.connect(
        host=db_host,
        port=int(db_port),
        user=db_user,
        password=db_password,
        db=db_name,
        charset=db_charset
    )

    # 创建 游标
    cur = conn.cursor()

    # 查询
    cur.execute(query_sql)

    # 获取 行数
    resultsRows = cur.rowcount

    # 关闭 连接
    cur.close()
    conn.close()

    return resultsRows

''' 将 结果，写入到 excel --不分国家'''
from openpyxl import Workbook
from openpyxl.styles import numbers
def write_excel(retenion_results, retenion_day):
    print("开始写入excel！")
    wb = Workbook()
    filename = '../retenion_datas_all.xlsx'
    ws1 = wb.active
    ws1.title = "retenion_all"

    ''' 写入 留存人数 信息 '''
    ws1.cell(column=1, row=1, value="注册日期")

    col = 2
    for d in retenion_day:
        val = str(d+1) + ' day'
        ws1.cell(column=col, row=1, value=val)
        col = col + 1

    col = 1
    row = 2
    for data in retenion_results:
        register_date = data['register_date']
        retenion_count = data['retenion_day']
        ws1.cell(column=col, row=row, value=register_date)
        col = col + 1
        for c in retenion_count:
            count = c['count']
            ws1.cell(column=col, row=row, value=count)
            col = col + 1
        row = row + 1
        col = 1

    ''' 写入 留存率 信息 '''
    row = row + 2  # 在 上面row的值 上，再加2，即：空2行
    col = 2
    ws1.cell(column=1, row=row, value="注册日期")
    for d in retenion_day:
        val = str(d+1) + ' day'
        ws1.cell(column=col, row=row, value=val)
        col = col + 1

    col = 1
    row = row + 1
    #min_row = row  # 制图需要，不要标题
    for data in retenion_results:
        register_date = data['register_date']
        retenion_count = data['retenion_day']
        register_player = retenion_count[0]['count']
        ws1.cell(column=col, row=row, value=register_date)
        col = col + 1
        for c in retenion_count:
            count = c['count']
            try:
                retenion_rate = float('%.4f' %(count / register_player))    # 百分比 运算
            except ZeroDivisionError:
                retenion_rate = 0
            ws1.cell(column=col, row=row, value=retenion_rate).number_format = numbers.FORMAT_PERCENTAGE_00   # 结果格式为：0.00%
            col = col + 1
        row = row + 1
        #max_col = col - 1
        col = 1
    #max_row = row - 1

    wb.save(filename)
    print("DONE!")
    return

''' 将 结果，写入到 excel --分国家'''
def write_country_excel(retenion_results, retenion_day, retenion_country):
    print("开始写入excel！")
    wb = Workbook()
    filename = '../retenion_datas_country.xlsx'
    for rc in retenion_country:
        title = "retenion_" + rc
        ws = wb.create_sheet(title=title)
        ws.cell(column=1, row=1, value="注册日期")

        col = 2
        for d in retenion_day:
            val = str(d + 1) + ' day'
            ws.cell(column=col, row=1, value=val)
            col = col + 1

        ''' 写入 留存人数 信息 '''
        col = 1
        row = 2
        for rr in retenion_results:
            if rr['country'] == rc:
                register_date = rr['register_date']
                retenion_count = rr['retenion_day']
                ws.cell(column=col, row=row, value=register_date)
                col = col + 1
                for c in retenion_count:
                    count = c['count']
                    ws.cell(column=col, row=row, value=count)
                    col = col + 1
                row = row + 1
                col = 1

        ''' 写入 留存率 信息 '''
        row = row + 2  # 在 上面row的值 上，再加2，即：空2行
        col = 2
        ws.cell(column=1, row=row, value="注册日期")
        for d in retenion_day:
            val = str(d + 1) + ' day'
            ws.cell(column=col, row=row, value=val)
            col = col + 1

        col = 1
        row = row + 1
        # min_row = row  # 制图需要，不要标题
        for rr in retenion_results:
            if rr['country'] == rc:
                register_date = rr['register_date']
                retenion_count = rr['retenion_day']
                register_player = retenion_count[0]['count']
                ws.cell(column=col, row=row, value=register_date)
                col = col + 1
                for c in retenion_count:
                    count = c['count']
                    try:
                        retenion_rate = float('%.4f' % (count / register_player))  # 百分比 运算
                    except ZeroDivisionError:
                        retenion_rate = 0
                    ws.cell(column=col, row=row,
                             value=retenion_rate).number_format = numbers.FORMAT_PERCENTAGE_00  # 结果格式为：0.00%
                    col = col + 1
                row = row + 1
                # max_col = col - 1
                col = 1
            # max_row = row - 1

    wb.save(filename)
    print("DONE!")
    return

if __name__ == '__main__':
    get_retention()
